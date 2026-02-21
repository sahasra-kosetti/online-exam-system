from flask import Flask, request, redirect, session, url_for
import time
import openpyxl
import os

app = Flask(__name__)
app.secret_key = "exam_system_secret"

# Check if running on Render
ON_RENDER = os.environ.get('RENDER', False)

if ON_RENDER:
    # On Render, use /tmp directory which is writable
    EXCEL_FILE = '/tmp/results.xlsx'
else:
    # Local development
    EXCEL_FILE = os.path.join(os.getcwd(), "results.xlsx")

EXAM_DURATION = 600   # 10 minutes

# CREATE EXCEL FILE
try:
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exam Results"
        ws.append([
            "Username", "Name", "Department", "Roll No",
            "Score", "Total", "Percentage", "Grade", "Result", "Time"
        ])
        wb.save(EXCEL_FILE)
except Exception as e:
    print(f"Warning: Could not create Excel file: {e}")
    # Continue anyway - the app can still work
    pass


# =====================================================
# SAVE RESULT
# =====================================================

def save_result(username, student, score, total, percent, grade, result):
    try:
        # Try to load existing workbook
        if os.path.exists(EXCEL_FILE):
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            # Create new workbook if doesn't exist
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Exam Results"
            ws.append([
                "Username", "Name", "Department", "Roll No",
                "Score", "Total", "Percentage", "Grade", "Result", "Time"
            ])
        
        # Add the result
        ws.append([
            username,
            student["name"],
            student["dept"],
            student["roll"],
            score,
            total,
            round(percent,2),
            grade,
            result,
            time.strftime("%Y-%m-%d %H:%M:%S")
        ])
        
        # Save the file
        wb.save(EXCEL_FILE)
        print(f"✓ Result saved for {username}")
        
    except Exception as e:
        # Print error but don't stop the app
        print(f"✗ Could not save result for {username}: {e}")
        # Optionally: Save to a backup text file
        try:
            with open('/tmp/backup_results.txt', 'a') as f:
                f.write(f"{username},{student['name']},{score},{total},{percent},{grade},{result},{time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        except:
            pass


# =====================================================
# STUDENTS DATA (MANUAL)
# =====================================================

students = {

"d_pravallika":{
    "password":"7894",
    "name":"D.Pravallika",
    "dept":"BCA Data Science",
    "roll":"2347390685"
},

"amani":{
    "password":"1235",
    "name":"A. Amani",
    "dept":"BCA Data Science",
    "roll":"BCA2025002"
},

"bhargavi":{
    "password":"1456",
    "name":"K.Bhargavi Satya Durga Devi ",
    "dept":"BCA Data Science",
    "roll":"2347390487"
},

"mahalakshmi":{
    "password":"7794",
    "name":" K.Mahalakshmi",
    "dept":"BCA Data Science",
    "roll":"2347390121"
},

"y_deepthi":{
    "password":"8964",
    "name":"Y.Deepthi Raajitha",
    "dept":"BCA Data Science",
    "roll":"2347390643"
},

"madhulika ":{
    "password":"7852",
    "name":"N.L.Madhulika ",
    "dept":"BCA Data Science",
    "roll":"2347390469"
},
   
"bhavya":{
    "password":"7410",
    "name":"K.Bhavya sri",
    "dept":"BCA Data Science",
    "roll":"2347390453"
},

"mrudula":{
    "password":"9856",
    "name":" P. Mrudula ",
    "dept":"BCA Data Science",
    "roll":" 2347390537"
},
"m_surekha":{
    "password":"5632",
    "name":" M.Surekha",
    "dept":"BCA Data Science",
    "roll":"2347390168"
},

"p_lalitha":{
    "password":"2587",
    "name":" P. Lalitha",
    "dept":"BCA Data Science",
    "roll":"2347390360"
},

"lini":{
    "password":"4758",
    "name":" D. Lini Hadassa",
    "dept":"BCA Data Science",
    "roll":"2347390616"
},
"venkatalakshmi":{
    "password":"6354",
    "name":"K. Venkata Lakshmi ",
    "dept":"BCA Data Science",
    "roll":"2347390581"
},

"nandinidevi":{
    "password":"4789",
    "name":" V. Nandini Devi",
    "dept":"BCA Data Science",
    "roll":" 2347390701"
},

"jnana_deepthi":{
    "password":"2587",
    "name":"K.Jnana Deepthi ",
    "dept":"BCA Data Science",
    "roll":"2347390592"
},
"b_reshma":{
    "password":"3214",
    "name":"B. Reshma",
    "dept":"BCA Data Science",
    "roll":" 2347390238"
},

"manga_charmila":{
    "password":"1645",
    "name":"L. Manga Charmila",
    "dept":"BCA Data Science",
    "roll":"2347390628"
},

"p_meghana":{
    "password":"8547",
    "name":"P.s.s.meghana ",
    "dept":"BCA Data Science",
    "roll":"2347390478"
},
"swarna":{
    "password":"9745",
    "name":"B. Swarna siva durga",
    "dept":"BCA Data Science",
    "roll":"2347390612"
},

"manogna":{
    "password":"6523",
    "name":"Alla Manogna ",
    "dept":"BCA Data Science",
    "roll":"2347390558"
},

"rajeswari":{
    "password":"6578",
    "name":"Sadanala Bhagyasri Raja Rajeswari Devi ",
    "dept":"BCA Data Science",
    "roll":"2347390700"
},
"sathvika":{
    "password":"9872",
    "name":"K.Sathvika",
    "dept":"BCA Data Science",
    "roll":"2347390532"
},

"saniya":{
    "password":"3645",
    "name":"Student Name 2",
    "dept":"BCA Data Science",
    "roll":"BCA2025002"
},

"teja":{
    "password":"6874",
    "name":"B. Teja Sri ",
    "dept":"BCA Data Science",
    "roll":"2347390138"
},

"suryalakshmi":{
    "password":"2035",
    "name":"Suryalakshmi.E",
    "dept":"BCA Data Science",
    "roll":"2347390477"
},

"j_meghana":{
    "password":"9547",
    "name":"JITTUGA MEGHANA ",
    "dept":"BCA Data Science",
    "roll":"2347390663"
},
"nissi_pravalika":{
    "password":"7894",
    "name":"BARLANKI NISSI PRAVALLIKA",
    "dept":"BCA Data Science",
    "roll":"2347390245"
},

"p_prasanna":{
    "password":"6841",
    "name":"P Naga Prasanna",
    "dept":"BCA Data Science",
    "roll":"2347390169"
},

"p_aparna":{
    "password":"6548",
    "name":"P.Aparna devi",
    "dept":"BCA Data Science",
    "roll":"2347390085"
},
"sunitha":{
    "password":"3258",
    "name":"A.Pranaya Sunitha ",
    "dept":"BCA Data Science",
    "roll":"2347390436"
},
"umadevi":{
    "password":"9547",
    "name":"K.Umadevi",
    "dept":"BCA Data Science",
    "roll":"2347390375"
},
"ashrita":{
    "password":"5874",
    "name":"N.Ashrita",
    "dept":"BCA Data Science",
    "roll":"2347390029"
},
"sailaja":{
    "password":"2587",
    "name":"A.Sailaja",
    "dept":"BCA Data Science",
    "roll":"2347390715"
},
"shalini":{
    "password":"7414",
    "name":"",
    "dept":"BCA Data Science",
    "roll":"2347392179"
},
"moksha":{
    "password":"6345",
    "name":"Ch. Moksha",
    "dept":"BCA Data Science",
    "roll":"2347390710"
},
"neeraja":{
    "password":"3574",
    "name":" V. Neeraja",
    "dept":"BCA Data Science",
    "roll":"2347390387"
},
"r_keerthi":{
    "password":"7913",
    "name":"R . Keerthi ",
    "dept":"BCA Data Science",
    "roll":"2347390242"
},
"sowmya":{
    "password":"5741",
    "name":"T. S. L. Sowmya",
    "dept":"BCA Data Science",
    "roll":"2347390024"
},
"charishma":{
    "password":"3247",
    "name":"Ch. Charishma",
    "dept":"BCA Data Science",
    "roll":"2347390186"
},

"nissi_tejaswini":{
    "password":"2147",
    "name":"G.Nissi Tejaswini",
    "dept":"BCA Data Science",
    "roll":" 2347390672"
},
"jayasri":{
    "password":"1047",
    "name":"P.Aparna devi",
    "dept":"BCA Data Science",
    "roll":"2347390329"
},
"b_jyothi":{
    "password":"6374",
    "name":" Boddu Nirmala Jyothi ",
    "dept":"BCA Data Science",
    "roll":" 2347390148"
},
"harika":{
    "password":"5471",
    "name":"Voleti Harika",
    "dept":"BCA Data Science",
    "roll":"2347390379"
},
"sudeeksha":{
    "password":"6474",
    "name":"N.M.S.Sudeeksha",
    "dept":"BCA Data Science",
    "roll":"2347390230"
},
"sahithi":{
    "password":"2471",
    "name":"K . Sahithi Ramani ",
    "dept":"BCA Data Science",
    "roll":"2347390533"
},
"a_mounika":{
    "password":"9707",
    "name":"Aari Mounika Devi",
    "dept":"BCA Data Science",
    "roll":"2347390517"
},
"prasanna":{
    "password":"3287",
    "name":"Eerlu Devi Krishna Prasanna ",
    "dept":"BSC Computer Science",
    "roll":"2347390312"
},
"d_lavanya":{
    "password":"6574",
    "name":"LAVANYA DOMMETI ",
    "dept":"BSC Computer Science",
    "roll":"2347390229"
},
"bindu":{
    "password":"1047",
    "name":"BINDU VIJAYA DURGA RAJESWARI MEDASANI",
    "dept":"BSC Computer Science",
    "roll":"2347390486"
},
"siva":{
    "password":"3017",
    "name":"SIVA SAINI YASALAPU",
    "dept":"BSC Computer Science",
    "roll":"2347390429"
},
"santhoshi":{
    "password":"9014",
    "name":"SANTHOSHI LAKSHMI GODAVARTHI",
    "dept":"BSC Computer Science",
    "roll":"2347390464"
},
"lalitha":{
    "password":"3501",
    "name":"LALITHA DEVI TEEDA ",
    "dept":"BSC Computer Science",
    "roll":"2347390020"
},
"durga":{
    "password":"6731",
    "name":"DURGA SADHANA GINJALA ",
    "dept":"BSC Computer Science",
    "roll":"2347390019"
},
"veera_mani":{
    "password":"9301",
    "name":"VEERA MANI GANDI",
    "dept":"BSC Computer Science",
    "roll":"2347390039"
},
"sahitya":{
    "password":"1478",
    "name":"MULUKURI SAHITYA DEEPIKA",
    "dept":"BSC Computer Science",
    "roll":"2347390366"
},
"tanuja":{
    "password":"6123",
    "name":"TANUJA CHOKKA",
    "dept":"BSC Computer Science",
    "roll":"2347390546"
},
"anusha":{
    "password":"3104",
    "name":"ANUSHA NAGIREDDY",
    "dept":"BSC Computer Science",
    "roll":"23473900251"
},
"nandeeswari":{
    "password":"3204",
    "name":"NANDEESWARI DODDI ",
    "dept":"BSC Computer Science",
    "roll":"2347390241"
},

"saranya":{
    "password":"9314",
    "name":"VEERA DURGA SARANYA AMARA ",
    "dept":"BSC Computer Science",
    "roll":"2347390034"
},

"sumana":{
    "password":"6475",
    "name":"JHANSI SUMANA KARRI",
    "dept":"BSC Computer Science",
    "roll":"2347390031"
},
"madhu_latha":{
    "password":"1532",
    "name":"DURGA MADHU LATHA SIRINGI",
    "dept":"BSC Computer Science",
    "roll":"2347390543"
},
"mounika":{
    "password":"4369",
    "name":"MOUNIKA ВОККА",
    "dept":"BSC Computer Science",
    "roll":"2347390223"
},

"yamini":{
    "password":"9547",
    "name":"YAMINI A",
    "dept":"BSC Computer Science",
    "roll":"2347390334"
},
"kavya":{
    "password":"0369",
    "name":"SAI DURGA KAVYA ANJALI PANDRADA",
    "dept":"BSC Computer Science",
    "roll":"2347390074"
},
"divya":{
    "password":"0234",
    "name":"DIVYA DURGA GANGA BHAVANI ACHANTA",
    "dept":"BSC Computer Science",
    "roll":"2347390289"
},
"swathi":{
    "password":"5024",
    "name":"SWATHI NATRA",
    "dept":"BSC Computer Science",
    "roll":"2347390239"
},
"surekha":{
    "password":"0478",
    "name":"LAKSHMI SUREKHA KANCHUMENL",
    "dept":"BSC Computer Science",
    "roll":"2347390100"
},
"satya_sai":{
    "password":"5789",
    "name":"VENKATA SATYA SAI LAKSHMI MATTAPARTHI",
    "dept":"BSC Computer Science",
    "roll":"2347390555"
},
"veera_divya":{
    "password":"2019",
    "name":"VEERA DIVYA EAGALA",
    "dept":"BSC Computer Science",
    "roll":"2347390447"
},
"syamala":{
    "password":"0358",
    "name":"SESHA SAI SYAMALA MALLISETTI",
    "dept":"BSC Computer Science",
    "roll":"2347390397"
},
"janaki":{
    "password":"9036",
    "name":"JANAKI CHAVALA",
    "dept":"BSC Computer Science",
    "roll":"2347390541"
},
"aparna":{
    "password":"6321",
    "name":" MV SATYA GANGA BHAVANI APARNA SEEKALA",
    "dept":"BSC Computer Science",
    "roll":"2347390596"
},

"reshma":{
    "password":"0136",
    "name":"RESHMA GUTHULA",
    "dept":"BSC Computer Science",
    "roll":"2347390038"
},
"m_kavya":{
    "password":"9370",
    "name":"SRI KAVYA MEDISETTI",
    "dept":"BSC Computer Science",
    "roll":"2347390094"
},
"y_swathi":{
    "password":"2002",
    "name":"VEERA SWATHI YALLA",
    "dept":"BSC Computer Science",
    "roll":"2347390205"
},

"k_vardhini":{
    "password":"3701",
    "name":"VENKATA SATYA VARDHINI KADAL",
    "dept":"BSC Computer Science",
    "roll":"2347390240"
},
"umamaheswari":{
    "password":"3012",
    "name":"UMAMAHESWARI GEDDADA",
    "dept":"BSC Computer Science",
    "roll":"2347390610"
},
"d_lalitha":{
    "password":"3214",
    "name":"LALITHA DESETTI",
    "dept":"BSC Computer Science",
    "roll":"2347390353"
},
"susmitha":{
    "password":"1532",
    "name":"V V T SUSMITHA APPANA",
    "dept":"BSC Computer Science",
    "roll":"2347390359"
},
"akshaya":{
    "password":"4369",
    "name":"AKSHAYA PULAVARTHY",
    "dept":"BSC Computer Science",
    "roll":"2347390369"
},

"lavanya":{
    "password":"0735",
    "name":"LAVANYA KARE",
    "dept":"BSC Computer Science",
    "roll":"2347390296"
},
"g_lalitha":{
    "password":"3258",
    "name":"LALITHA GUTHULA",
    "dept":"BSC Computer Science",
    "roll":"2347390424"
},
"pravallika":{
    "password":"2005",
    "name":"PRAVALLIKA BALIJI",
    "dept":"BSC Computer Science",
    "roll":"2347390363"
},
"jyothi":{
    "password":"1478",
    "name":"JYOTHI NIHARIKA JALLIGAMPALA",
    "dept":"BSC Computer Science",
    "roll":"2347390471"
},
"madhavi":{
    "password":"3259",
    "name":"MADHAVI SURYA CHANDRIKA MEDA",
    "dept":"BSC Computer Science",
    "roll":"2347390284"
},
"yashmitha":{
    "password":"5963",
    "name":"YASHMITHA JYOTHIKA BHAVANI BURRA",
    "dept":"BSC Computer Science",
    "roll":"2347390327"
},
"jansi":{
    "password":"0796",
    "name":"JANSI SRI SATYA ΜΑΤΤΑ",
    "dept":"BSC Computer Science",
    "roll":"2347390542"
},
"snehita":{
    "password":"3896",
    "name":"SNEHITA REDDY PENUMALLU",
    "dept":"BSC Computer Science",
    "roll":"2347390400"
},
"k_bhuvaneswari":{
    "password":"0199",
    "name":"BHUVANESWARI KARELLA",
    "dept":"BSC Computer Science",
    "roll":"2347390597"
},
"amulya":{
    "password":"5263",
    "name":"SRI AMULYA KALIDINDI",
    "dept":"BSC Computer Science",
    "roll":"2347390086"
},
"uma_pavani":{
    "password":"3257",
    "name":"NAGA LAKSHMI UMA PAVANI KANKATALA",
    "dept":"BSC Computer Science",
    "roll":"2347390261"
},
"likitha":{
    "password":"9314",
    "name":"SRI ANANTHA LAKSHMI LIKITHA POLAVARAPU",
    "dept":"BSC Computer Science",
    "roll":"2347390457"
},
"ramya":{
    "password":"5860",
    "name":"JAYA RAMYA NAGA JYOTHIKA KORN",
    "dept":"BSC Computer Science",
    "roll":"2347390341"
},
"bhuvaneswari":{
    "password":"1470",
    "name":"BHUVANESWARI BORA",
    "dept":"BSC Computer Science",
    "roll":"2347390540"
},
"satyavathi":{
    "password":"9874",
    "name":"SATYAVATHI YEDLA",
    "dept":"BSC Computer Science",
    "roll":"2347390613"
},
"padma_priya":{
    "password":"2580",
    "name":"PADMA PRIYA SRI MAHA LAKSHMI TAMALAMPUDI",
    "dept":"BSC Computer Science",
    "roll":"2347390516"
},
"vijaya":{
    "password":"3690",
    "name":"VIJAYA SAI SARANYA CHINTHAKULA",
    "dept":"BSC Computer Science",
    "roll":"2347390275"
},
"k_mounika":{
    "password":"3796",
    "name":"MOUNIKA SURYA LAKSHMI KOSURI",
    "dept":"BSC Computer Science",
    "roll":"2347390159"
},
"uma_devi":{
    "password":"0845",
    "name":"Uma devi M",
    "dept":"BSC Computer Science",
    "roll":"2347390602"
},
"pushpanjali":{
    "password":"1428",
    "name":"PUSHPANJALI MEDAPATI",
    "dept":"BSC Computer Science",
    "roll":"2347390117"
},
"sandhya":{
    "password":"6230",
    "name":"SANDHYA KONDEPUD",
    "dept":"BSC Computer Science",
    "roll":"2347390152"
},
"m_aparna":{
    "password":"7063",
    "name":"APARNA MARISE",
    "dept":"BSC Computer Science",
    "roll":"2347390315"
},
"revathi":{
    "password":"1733",
    "name":"REVATHI ADI",
    "dept":"BSC Computer Science",
    "roll":"2347390495"
},
"supriya":{
    "password":"3589",
    "name":"BALA KAMESHWARI SUPRIYA VELUGUBANT",
    "dept":"BSC Computer Science",
    "roll":"2347390438"
},
"roopa":{
    "password":"8017",
    "name":"ROOPA VANI PEKETI",
    "dept":"BSC Computer Science",
    "roll":"2347390696"
},
"devi":{
    "password":"3275",
    "name":"DEVI PABBINEEDI",
    "dept":"BSC Computer Science",
    "roll":"2347390662"
},

    
# continue till student120

"sahasra":{
    "password":"1234",
    "name":"Sahasra Kosetti",
    "dept":"BCA Data Science",
    "roll":"2347390190"
}

}


# =====================================================
# QUESTIONS
# =====================================================

questions = {

    1: {
        "text": "Which of the following protocols is used for secure communication over the Internet?",
        "options": ["HTTP", "FTP", "HTTPS", "SMTP"]
    },

    2: {
        "text": "What is the time complexity of binary search algorithm?",
        "options": ["O(n)", "O(log n)", "O(n log n)", "O(1)"]
    },

    3: {
        "text": "Which layer of OSI model is responsible for logical addressing?",
        "options": ["Transport Layer", "Network Layer", "Session Layer", "Data Link Layer"]
    },

    4: {
        "text": "Which data structure uses FIFO principle?",
        "options": ["Stack", "Queue", "Tree", "Graph"]
    },

    5: {
        "text": "Which SQL command is used to remove a table permanently?",
        "options": ["DELETE", "REMOVE", "DROP", "CLEAR"]
    },

    6: {
        "text": "Which of the following is NOT an operating system?",
        "options": ["Linux", "Windows", "Oracle", "MacOS"]
    },

    7: {
        "text": "Which normal form removes transitive dependency?",
        "options": ["1NF", "2NF", "3NF", "BCNF"]
    },

    8: {
        "text": "Which protocol is used to send emails?",
        "options": ["FTP", "SMTP", "HTTP", "SNMP"]
    },

    9: {
        "text": "Which sorting algorithm has best average time complexity?",
        "options": ["Bubble Sort", "Selection Sort", "Merge Sort", "Insertion Sort"]
    },

    10: {
        "text": "Which of the following is used for version control?",
        "options": ["Git", "Python", "MySQL", "HTML"]
    },

    11: {
        "text": "Which memory is fastest?",
        "options": ["RAM", "ROM", "Cache", "Hard Disk"]
    },

    12: {
        "text": "Which symbol is used for single line comment in Python?",
        "options": ["//", "#", "/* */", "--"]
    },

    13: {
        "text": "Which protocol is used for file transfer?",
        "options": ["FTP", "HTTP", "SMTP", "TCP"]
    },

    14: {
        "text": "Which data structure uses LIFO principle?",
        "options": ["Queue", "Stack", "Array", "Tree"]
    },

    15: {
        "text": "Which key is used to uniquely identify a record in database?",
        "options": ["Foreign Key", "Primary Key", "Candidate Key", "Alternate Key"]
    },
    16: {
        "text": "Which of the following is volatile memory?",
        "options": ["ROM", "Hard Disk", "RAM", "CD-ROM"]
    },

    17: {
        "text": "Which protocol is used to access web pages?",
        "options": ["HTTP", "FTP", "SMTP", "POP"]
    },

    18: {
        "text": "Which of the following is not a programming language?",
        "options": ["Python", "Java", "HTML", "Windows"]
    },

    19: {
        "text": "Which data structure is used in recursion?",
        "options": ["Queue", "Stack", "Tree", "Graph"]
    },

    20: {
        "text": "Which of the following is example of system software?",
        "options": ["MS Word", "Windows OS", "Chrome", "Photoshop"]
    },

    21: {
        "text": "Which SQL clause is used to filter records?",
        "options": ["WHERE", "ORDER", "GROUP", "FILTER"]
    },

    22: {
        "text": "Which is brain of computer?",
        "options": ["RAM", "CPU", "Hard Disk", "Monitor"]
    },

    23: {
        "text": "Which topology uses central hub?",
        "options": ["Ring", "Bus", "Star", "Mesh"]
    },

    24: {
        "text": "Which of the following is NoSQL database?",
        "options": ["MySQL", "Oracle", "MongoDB", "SQL Server"]
    },

    25: {
        "text": "Which keyword is used to create function in Python?",
        "options": ["function", "def", "create", "fun"]
    },

    26: {
        "text": "Which of the following is output device?",
        "options": ["Keyboard", "Mouse", "Printer", "Scanner"]
    },

    27: {
        "text": "Which layer ensures reliable data transfer?",
        "options": ["Transport Layer", "Application Layer", "Physical Layer", "Session Layer"]
    },

    28: {
        "text": "Which operator is used for exponent in Python?",
        "options": ["^", "**", "%", "//"]
    },

    29: {
        "text": "Which is default port of HTTP?",
        "options": ["80", "443", "21", "25"]
    },

    30: {
        "text": "Which is example of interpreted language?",
        "options": ["C", "C++", "Python", "Assembly"]
    } 

}


# =====================================================
# ANSWERS
# =====================================================

answer_key = {

    1: "HTTPS",
    2: "O(log n)",
    3: "Network Layer",
    4: "Queue",
    5: "DROP",
    6: "Oracle",
    7: "3NF",
    8: "SMTP",
    9: "Merge Sort",
    10: "Git",
    11: "Cache",
    12: "#",
    13: "FTP",
    14: "Stack",
    15: "Primary Key",
    16: "RAM",
    17: "HTTP",
    18: "Windows",
    19: "Stack",
    20: "Windows OS",
    21: "WHERE",
    22: "CPU",
    23: "Star",
    24: "MongoDB",
    25: "def",
    26: "Printer",
    27: "Transport Layer",
    28: "**",
    29: "80",
    30: "Python"

}


# =====================================================
# GRADE FUNCTION
# =====================================================

def grade(p):

    if p>=90:return "A+"
    elif p>=75:return "A"
    elif p>=60:return "B"
    elif p>=50:return "C"
    else:return "F"

# =====================================================
# LOGOUT
# =====================================================

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")



# =====================================================
# PROFESSIONAL CSS + TIMER SCRIPT
# =====================================================

css = """

<style>

body{
margin:0;
font-family:Segoe UI;
background: linear-gradient(90deg, #F6CFBE, #B9DCF2);
}

/* LOGIN */

.login{

width:400px;
margin:auto;
margin-top:120px;
background:white;
padding:40px;
border-radius:12px;
}

.center{text-align:center;}

input{

width:100%;
padding:12px;
margin-top:10px;
margin-bottom:20px;

}
.question input[type="radio"]{
width:auto;
margin-right:8px;
}


button{

padding:12px 40px;
background:#ff512f;
color:white;
border:none;
border-radius:25px;

}


/* EXAM */

.exam{

display:flex;
justify-content:center;
gap:40px;
margin-top:20px;

}

.logo img{

width:140px;

}

.timer{

color:white;
font-size:28px;

}

.paper{

background:white;
padding:30px;
border-radius:12px;
width:700px;

}

.question{

margin-top:15px;

}

.result{

width:400px;
margin:auto;
margin-top:100px;
background:white;
padding:30px;
border-radius:12px;
text-align:center;

}

.pass{color:green;font-size:22px;}
.fail{color:red;font-size:22px;}
.grade{color:blue;font-size:20px;}

</style>


<script>

function startTimer(duration){

let timer=duration;

setInterval(function(){

let m=parseInt(timer/60)
let s=parseInt(timer%60)

m=m<10?"0"+m:m
s=s<10?"0"+s:s

document.getElementById("timer").innerHTML=m+":"+s

if(timer<=0){

document.getElementById("examForm").submit()

}

timer--

},1000)

}

</script>

"""


# =====================================================
# LOGIN PAGE
# =====================================================

@app.route("/",methods=["GET","POST"])
def login():

    if request.method=="POST":

        u=request.form["username"].strip()
        p=request.form["password"].strip()

        if u in students and students[u]["password"]==p:

            session["user"]=u
            session["start"]=time.time()

            return redirect("/exam")

    return css+"""

<div class="login">

<h2 class="center">Student Login</h2>

<form method="post">

Username
<input name="username" required>

Password
<input type="password" name="password" required>

<div class="center">
<button>Login</button>
</div>

</form>

</div>

"""


# =====================================================
# EXAM PAGE
# =====================================================

@app.route("/exam",methods=["GET","POST"])
def exam():

    if "user" not in session:
        return redirect("/")

    student=students[session["user"]]

    elapsed=time.time()-session["start"]

    remaining=int(EXAM_DURATION-elapsed)

    if remaining<=0:

        return redirect("/submit")

    if request.method=="POST":

        score=0

        for q in questions:

            if request.form.get(f"q{q}")==answer_key[q]:
                score+=1

        total=len(questions)

        percent=(score/total)*100

        g=grade(percent)

        result="PASS" if percent>=50 else "FAIL"

        save_result(session["user"],student,score,total,percent,g,result)

        return css+f"""

<div class="result">

<h2>Exam Result</h2>

Name: {student['name']}<br><br>
Dept: {student['dept']}<br><br>
Roll: {student['roll']}<br><br>

Score: {score}/{total}<br><br>

Percentage: {percent:.2f}%<br><br>

<div class="grade">Grade: {g}</div><br>

<div class="{'pass' if result=='PASS' else 'fail'}">
{result}
</div>

<br><br>

<a href="/logout">
<button>Logout</button>
</a>

</div>

"""



    html=css+f"""

<body onload="startTimer({remaining})">

<div class="exam">

<div class="logo">

<img src="/static/college_logo.png">

</div>

<div class="paper">

<h2 class="center">Question Paper</h2>

<form method="post" id="examForm">

"""


    for q,qdata in questions.items():

        html+=f"<div class='question'><b>Q{q}. {qdata['text']}</b><br>"

        for opt in qdata["options"]:

            html+=f"""
<input type="radio" name="q{q}" value="{opt}" required>
{opt}<br>
"""

        html+="</div>"


    html+="""


<br>

<div class="center">
<button>Submit Exam</button>
</div>

</form>

</div>

<div class="timer">

Time Left<br>

<span id="timer">10:00</span>

</div>

</div>

</body>

"""

    return html


# =====================================================
# RUN
# =====================================================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
